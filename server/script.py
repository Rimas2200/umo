# -*- coding: utf-8 -*-
import mysql.connector  # pip install mysql-connector-python
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
import os
import sys

if len(sys.argv) != 4:
    print("Неправильное количество аргументов")
    sys.exit(1)

faculty = sys.argv[1].lower()
semester = sys.argv[2]
academic_year = sys.argv[3]

bold_font = Font(bold=True)
center_alignment = Alignment(horizontal='center', vertical='center')


def format_cell_center(sheet, cell, text):
    sheet[cell] = text
    sheet[cell].font = bold_font
    sheet[cell].alignment = center_alignment


def format_cell(sheet, cell, text):
    sheet[cell] = text
    sheet[cell].font = bold_font


def sort_key(name):
    match = re.match(r"([А-Яа-яA-Za-z\-]+)-(\d+)", name)
    if match:
        letter_part = match.group(1)
        number_part = int(match.group(2))
        return letter_part, number_part
    return name, 0


def get_course(name):
    match = re.match(r"([А-Яа-яA-Za-z\-]+)-(\d+)", name)
    if match:
        number_part = int(match.group(2))
        course = number_part // 100
        return course
    return 0


# faculty = os.getenv('FACULTY').lower()
# semester = os.getenv('SEMESTER')
# academic_year = os.getenv('ACADEMIC_YEAR')
# faculty = "Математический факультет"
# semester = "3"
# academic_year = "2024 - 2025"
array = ["Понедельник", "Вторник", "Среда", "Четверг", "Пятница", "Суббота"]
merged_pairs_by_direction_course_day = {}
mysql_connection = mysql.connector.connect(
    host='127.0.0.1',
    user='root',
    password='',
    database='umo'
)
cursor = mysql_connection.cursor()

group_array = set()
cursor.execute("""SELECT name FROM group_name WHERE direction_abbreviation IN 
                            (SELECT direction_abbreviation FROM direction WHERE LOWER(faculty) = %s)""", (faculty,))
rows = cursor.fetchall()

group_array = [row[0] for row in rows]

groups_by_direction_and_course = defaultdict(list)

for group in group_array:
    match = re.match(r"([А-Яа-яA-Za-z\-]+)-(\d+)", group)
    if match:
        direction = match.group(1)
        course = get_course(group)
        groups_by_direction_and_course[(direction, course)].append(group)

for key in groups_by_direction_and_course:
    groups_by_direction_and_course[key] = sorted(groups_by_direction_and_course[key], key=sort_key)

wb = Workbook()
wb.remove(wb.active)

for course in range(1, 7):
    for direction in sorted(set([key[0] for key in groups_by_direction_and_course.keys()])):
        schedule_data = []
        if (direction, course) in groups_by_direction_and_course:
            sheet_name = " ".join(groups_by_direction_and_course[(direction, course)])
            sheet_name = sheet_name[:31]
            ws = wb.create_sheet(title=sheet_name)

           #заполнение файла
            format_cell(ws, 'J1', "УТВЕРЖДАЮ")
            format_cell(ws, 'J3', "ПРОРЕКТОР ПО УЧЕБНОЙ РАБОТЕ")
            format_cell(ws, 'J4', 'ФГБОУ ВО "ЧелГУ"          САЛАМАТОВ А.А.')
            format_cell(ws, 'J6', '"_______" __________________________2024 г.')

            for row in range(9, 13):
                ws.merge_cells(f'A{row}:J{row}')

            format_cell_center(ws, 'A9', "РАСПИСАНИЕ УЧЕБНЫХ ЗАНЯТИЙ")
            format_cell_center(ws, 'A10', f"на {semester} семестр {academic_year} учебный год")
            format_cell_center(ws, 'A11', faculty)
            format_cell_center(ws, 'A12', "Форма обучения ОЧНАЯ")

            start_col = 5  # нужно будет определить, сколько максимум подгрупп
            start_row = 15
            step = 2
            for i, group in enumerate(groups_by_direction_and_course[(direction, course)]):
                col_index = start_col + i * step
                col_letter = chr(col_index + 64)
                next_col_letter = chr(col_index + 65)
                cell = f"{col_letter}{start_row}"
                next_cell = f"{next_col_letter}{start_row}"
                ws.merge_cells(f"{cell}:{next_cell}")
                ws[cell] = group
                ws[cell].alignment = Alignment(horizontal='center', vertical='center')

                ws.column_dimensions[col_letter].width = max(len(group), 35)
                ws.column_dimensions[next_col_letter].width = max(len(group), 35)

                for day in array:
                    cursor.execute("""SELECT pair_name FROM schedule WHERE group_name = %s AND day_of_the_week = %s""", (group, day))
                    rows = cursor.fetchall()
                    unique_pairs = set()
                    for row in rows:
                        unique_pairs.add(row[0])
                    sorted_pairs = ', '.join(sorted(unique_pairs))
                    existing_entry = next((entry for entry in schedule_data if entry[0] == direction and entry[1] == course and entry[2] == day), None)

                    if existing_entry:
                        index = schedule_data.index(existing_entry)
                        combined_pairs = set(existing_entry[3].split(', '))
                        combined_pairs.update(unique_pairs)
                        schedule_data[index] = (existing_entry[0], existing_entry[1], existing_entry[2], ', '.join(sorted(combined_pairs)))
                    else:
                        schedule_data.append((direction, course, day, sorted_pairs))

            start_row_pair = 16
            start_row_pair_add = 16
            start_col_pair = 4
            step_pair = 1
            current_row = start_row_pair
            previous_day = None
            start_day_row = start_row_pair

            for entry in schedule_data:
                current_day = entry[2]
                if previous_day is not None and previous_day != current_day:
                    ws.merge_cells(f"C{start_day_row}:C{current_row - 1}")
                    cell = f"C{start_day_row}"
                    ws[cell] = previous_day
                    ws[cell].alignment = Alignment(text_rotation=90, vertical='center', horizontal='center')
                    start_day_row = current_row
                pairs = entry[3].split(', ')
                for pair in pairs:
                    cell = f"D{current_row}"
                    ws[cell] = f"{pair}"
                    current_row += 4

                previous_day = current_day
            if previous_day is not None:
                ws.merge_cells(f"C{start_day_row}:C{current_row - 1}")
                cell = f"C{start_day_row}"
                ws[cell] = previous_day
                ws[cell].alignment = Alignment(text_rotation=90, vertical='center', horizontal='center')

            group_col = 5
            group_row = 15
            pair_col = 4
            pair_row = 16
            day_col = 3
            day_row = 16
            n = 1
            day = 'День'
            schedule_col = 5
            schedule_row = 16

            while True:
                group = ws.cell(row=group_row, column=group_col).value

                if group is None:
                    # group_col = 5
                    # group_row = 15
                    # pair_row = 16
                    # day_row = 16

                    break
                while True:
                    if ws.cell(row=day_row, column=day_col).value:
                        day = ws.cell(row=day_row, column=day_col).value
                    pair = ws.cell(row=pair_row, column=pair_col).value
                    if pair is None:
                        break
                    cursor.execute("""SELECT * FROM schedule WHERE group_name = %s AND day_of_the_week = %s AND pair_name = %s""", (group, day, pair))
                    rows = cursor.fetchall()
                    if rows:
                        schedule_col = group_col
                        schedule_row = pair_row
                        for row in rows:  # row[8] - группа row[7] - неделя
                            column_letter = get_column_letter(schedule_col)
                            try:
                                if row[8] == '1' and row[7] == '1':  # 1 Неделя 1 Группа
                                    cell = f"{column_letter}{schedule_row}"
                                    ws[cell] = f"1Н {row[1]}"
                                    ws[cell].alignment = Alignment(horizontal='center')
                                    schedule_row += 1

                                    cell = f"{column_letter}{schedule_row}"
                                    ws[cell] = f"{row[5]}, {row[2]}"
                                    ws[cell].alignment = Alignment(horizontal='center')

                                elif row[8] == '2' and row[7] == '1':  # 1 Неделя 2 Группа
                                    schedule_col += 1
                                    column_letter = get_column_letter(schedule_col)
                                    cell = f"{column_letter}{schedule_row}"
                                    ws[cell] = f"1Н {row[1]}"
                                    ws[cell].alignment = Alignment(horizontal='center')
                                    schedule_row += 1

                                    cell = f"{column_letter}{schedule_row}"
                                    ws[cell] = f"{row[5]}, {row[2]}"
                                    ws[cell].alignment = Alignment(horizontal='center')

                                elif row[8] == '1' and row[7] == '2':  # 2 Неделя 1 Группа
                                    schedule_row += 1
                                    column_letter = get_column_letter(schedule_col)
                                    cell = f"{column_letter}{schedule_row}"
                                    ws[cell] = f"2Н {row[1]}"
                                    ws[cell].alignment = Alignment(horizontal='center')
                                    schedule_row += 1

                                    cell = f"{column_letter}{schedule_row}"
                                    ws[cell] = f"{row[5]}, {row[2]}"
                                    ws[cell].alignment = Alignment(horizontal='center')

                                elif row[8] == '2' and row[7] == '2':  # 2 Неделя 2 Группа
                                    schedule_col += 1
                                    column_letter = get_column_letter(schedule_col)
                                    schedule_row += 1
                                    cell = f"{column_letter}{schedule_row}"
                                    ws[cell] = f"2Н {row[1]}"
                                    ws[cell].alignment = Alignment(horizontal='center')
                                    schedule_row += 1
                                    cell = f"{column_letter}{schedule_row}"
                                    ws[cell] = f"{row[5]}, {row[2]}"
                                    ws[cell].alignment = Alignment(horizontal='center')

                                if row[8] == 'не определена' and row[7] == '1':  # 1 Неделя Обе группы
                                    schedule_row += 1
                                    column_letter = get_column_letter(schedule_col)
                                    next_column_letter = get_column_letter(schedule_col + 1)
                                    ws.merge_cells(f"{column_letter}{schedule_row}:{next_column_letter}{schedule_row}")
                                    ws[f"{column_letter}{schedule_row}"] = f"1Н {row[1]}"
                                    ws[f"{column_letter}{schedule_row}"].alignment = Alignment(horizontal='center')
                                    schedule_row += 1

                                    ws.merge_cells(f"{column_letter}{schedule_row}:{next_column_letter}{schedule_row}")
                                    ws[f"{column_letter}{schedule_row}"] = f"{row[5]}, {row[2]}"
                                    ws[f"{column_letter}{schedule_row}"].alignment = Alignment(horizontal='center')

                                elif row[8] == 'не определена' and row[7] == '2':  # 2 Неделя Обе группы
                                    schedule_row += 2
                                    column_letter = get_column_letter(schedule_col)
                                    next_column_letter = get_column_letter(schedule_col + 1)
                                    ws.merge_cells(f"{column_letter}{schedule_row}:{next_column_letter}{schedule_row}")
                                    ws[f"{column_letter}{schedule_row}"] = f"2Н {row[1]}"
                                    ws[f"{column_letter}{schedule_row}"].alignment = Alignment(horizontal='center')
                                    schedule_row += 1

                                    ws.merge_cells(f"{column_letter}{schedule_row}:{next_column_letter}{schedule_row}")
                                    ws[f"{column_letter}{schedule_row}"] = f"{row[5]}, {row[2]}"
                                    ws[f"{column_letter}{schedule_row}"].alignment = Alignment(horizontal='center')

                                elif row[8] == 'не определена' and row[7] == 'все':  # Все недели Обе группы
                                    schedule_row += 1
                                    column_letter = get_column_letter(schedule_col)
                                    next_column_letter = get_column_letter(schedule_col + 1)
                                    ws.merge_cells(f"{column_letter}{schedule_row}:{next_column_letter}{schedule_row}")
                                    ws[f"{column_letter}{schedule_row}"] = f"{row[1]}"
                                    ws[f"{column_letter}{schedule_row}"].alignment = Alignment(horizontal='center')
                                    schedule_row += 1

                                    ws.merge_cells(f"{column_letter}{schedule_row}:{next_column_letter}{schedule_row}")
                                    ws[f"{column_letter}{schedule_row}"] = f"{row[5]}, {row[2]}"
                                    ws[f"{column_letter}{schedule_row}"].alignment = Alignment(horizontal='center')
                            except: pass
                    pair_row += 4
                    day_row = pair_row

                group_col += 2
                n = 1
                pair_row = 16
                day_row = 16

file_name = f"{faculty}.xlsx"
wb.save(file_name)
print(f"Файл '{file_name}' успешно создан.")
