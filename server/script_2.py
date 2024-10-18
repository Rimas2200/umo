# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter
import mysql.connector   #pip install mysql-connector-python
import os
import sys

array_day = ('Понедельник', 'Вторник', 'Среда', 'Четверг', 'Пятница', 'Суббота')
wb = openpyxl.Workbook()
sheet = wb['Sheet']
sheet["B1"] = "семестр"
sheet["D1"] = "учебный год"
sheet.merge_cells('A2:B2')
sheet["A2"] = "№___"
sheet.merge_cells('A3:B3')
sheet["A3"] = "уч.кор"
classroom = "A1"
k = 3
kk = 3
classroom_row = 3
classroom_column = 3
n = 0
for i in range(len(array_day)):
    for j in range(1, 9):
        sheet.cell(row=k + j + n, column=2).value = str(j) + " 1Н"
        n += 1
        sheet.cell(row=k + j + n, column=2).value = str(j) + " 2Н"
    sheet.merge_cells(start_row=kk+1, start_column=1, end_row=kk+16, end_column=1)  # Объединение ячеек от A3 до B4
    sheet.cell(row=kk + 1, column=1).value = array_day[i]
    k+=8
    kk+=16
y = 3
mysql_connection = mysql.connector.connect(
    host='127.0.0.1',
    user='root',
    password='',
    database='umo'
)
lists = []
cursor = mysql_connection.cursor()
for day in range(len(array_day)):
    cursor.execute('''SELECT * FROM schedule''')
    
    results = cursor.fetchall()
    for row in results:
        lists.append(row[2])
    unique_elements = list(set(lists))
    filtered_elements = unique_elements
    # filtered_elements = [element for element in unique_elements if element]
    # filtered_elements = [element for element in filtered_elements if len(element) <= 12]
    for classroom in filtered_elements:
        cursor.execute('''SELECT * FROM schedule WHERE classroom = %s AND day_of_the_week = %s''', (classroom,str(array_day[day])))
        results = cursor.fetchall()
        for rows in results:
            print(rows)
            if rows[7] == "все" and rows[2] != ' ' and rows[4] != 'None':
                sheet.cell(row=classroom_row+(int(rows[4])*2)-1, column=classroom_column).value = rows[3]
                sheet.cell(row=classroom_row+int(rows[4])*2, column=classroom_column).value = rows[3]
            if rows[7] == "1" and rows[2] != ' ' and rows[4] != 'None':
                sheet.cell(row=classroom_row + int(rows[4])*2-1, column=classroom_column).value = rows[3]
            if rows[7] == "2" and rows[2] != ' ' and rows[4] != 'None':
                sheet.cell(row=classroom_row + int(rows[4])*2, column=classroom_column).value = rows[3]
        sheet.cell(row=2, column=classroom_column).value = str(classroom)
        column_width = 18
        column_letter = get_column_letter(classroom_column)
        column_dimensions = sheet.column_dimensions[column_letter]
        column_dimensions.width = column_width
        classroom_column += 1
        classroom_row = y
    y += 16
    classroom_column = 3
    day+=1
wb.save("Шахматка.xlsx")
mysql_connection.commit()

cursor.close()
mysql_connection.close()